from openpyxl import load_workbook
import os
import sys
import datetime
import xml.etree.ElementTree as ET

def extract_excel_to_xml(excel_file, output_xml):
    try:
        #load workbook and active sheet to work with excel file
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
    except Exception as e:
        print(f"Error loading Excel file {excel_file}: {e}")
        return

    try:
        #root element for the XML
        root = ET.Element("import")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                #create a node element
                node = ET.SubElement(root, "node", action="update", type="document")

                #add location as a sub-element
                location = ET.SubElement(node, "location")
                location.text = f"Enterprise:Engineering:Legacy Records:{row[9]}:{row[1]}:{row[0]}"

                #add category with attributes as sub-elements
                category = ET.SubElement(node, "category", name="Content Server Categories:HC Engineering:Engineering Legacy Records")
                attributes = {
                    "Project Name": row[1],
                    "Permit Number": row[2],
                    "Box Bar Code": row[3],
                    "GIS FeatureID": row[4],
                    "Address": row[5],
                    "Description": row[6],
                    "Project Type": row[9].strip().title() if row[9] else "",
                    "Document Date": datetime.date.strftime(row[10], '%Y%m%d') if row[10] else "",
                    "Document Type": row[11].strip().title() if row[11] else "",
                    "Job Number": row[12],
                    "Precinct": row[13],
                    "PIN": row[14],
                    "Project Limits": row[15],
                    "Keywords": row[16],
                    "Metadata Reviewed": row[17]
                }

                for key, value in attributes.items():
                    attribute = ET.SubElement(category, "attribute", name=key)
                    attribute.text = str(value) if value else ""

        #write the constructed XML tree to a file
        tree = ET.ElementTree(root)
        tree.write(output_xml, encoding="utf-8", xml_declaration=True)
        print(f"XML file saved: {output_xml}")

    except Exception as e:
        print(f"Error generating XML for {excel_file}: {e}")

def process_directory(input_dir, output_dir):
    file_count = 0

    try:
        for subdir_name in os.listdir(input_dir):
            subdir_path = os.path.join(input_dir, subdir_name)

            if os.path.isdir(subdir_path):
                excel_files = [f for f in os.listdir(subdir_path) if f.endswith('Full Index.xlsx')]
                
                if excel_files:
                    excel_file = os.path.join(subdir_path, excel_files[0])
                    output_subdir = os.path.join(output_dir, subdir_name)
                    os.makedirs(output_subdir, exist_ok=True)
                    output_xml = os.path.join(output_subdir, f"{subdir_name}_{os.path.splitext(excel_files[0])[0]}.xml")

                    try:
                        extract_excel_to_xml(excel_file, output_xml)
                        file_count += 1
                    except Exception as e:
                        print(f"Failed to process {excel_file}: {e}")
    except Exception as e:
        print(f"Error processing directory {input_dir}: {e}")

    print(f"Total files processed: {file_count}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py <input_directory> <output_directory>")
        sys.exit(1)

    input_dir = sys.argv[1]
    output_dir = sys.argv[2]

    try:
        os.makedirs(output_dir, exist_ok=True)
        process_directory(input_dir, output_dir)
    except Exception as e:
        print(f"An error occurred: {e}")
